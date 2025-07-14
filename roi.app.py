import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from openpyxl import load_workbook

st.set_page_config(page_title="Commonwealth Cares Deployment Model", layout="wide")

st.title("📊 Commonwealth Cares Deployment & ROI Calculator")

st.markdown("This tool allows you to visualize staffing deployment assumptions and shift-based ROI scenarios.")

uploaded_file = st.file_uploader("Upload your Excel File", type=["xlsx"])

if uploaded_file:
    tab1, tab2 = st.tabs(["Deployment Visualizer", "Shift-Based ROI Calculator"])

    with tab1:
                    wb = load_workbook(uploaded_file, data_only=True)
    sheet = wb['Combo- Select & Float Pool']

    editable_values = []
    for row in sheet.iter_rows():
        for idx, cell in enumerate(row):
            if isinstance(cell.value, (int, float)):
                if cell.font and cell.font.color and cell.font.color.type == 'rgb':
                    if cell.font.color.rgb.upper().startswith('FFFF0000'):
                        description = None
                        for prev_cell in reversed(row[:idx]):
                            if isinstance(prev_cell.value, str) and prev_cell.value.strip() != "":
                                description = prev_cell.value.strip()
                                break
                        if description:
                            editable_values.append({'description': description, 'value': cell.value, 'cell': cell.coordinate})

    st.sidebar.header("🔧 Adjust Model Inputs")
    input_values = {}

    with st.sidebar.expander("📌 Permanent Staffing Inputs"):
        st.markdown("**Permanent staffing ramp-up assumptions.**")
        for item in editable_values:
            if item['cell'] in ['B21', 'B22', 'C17']:
                label = f"{item['description']} ({item['cell']})"
                input_values[item['cell']] = st.slider(
                    label,
                    key=f"{item['cell']}_{label}",
                    min_value=0,
                    max_value=int(item['value'] * 4),
                    value=int(item['value'])
                )

    with st.sidebar.expander("📌 Float Pool Inputs"):
        st.markdown("**Float Pool deployment assumptions.**")
        for item in editable_values:
            if item['cell'] in ['C17', 'B27', 'B26']:
                label = f"{item['description']} ({item['cell']})"
                input_values[item['cell']] = st.slider(
                    label,
                    min_value=0,
                    max_value=int(item['value'] * 4),
                    value=int(item['value'])
                )

    with st.sidebar.expander("📌 VISTA Locums Inputs"):
        st.markdown("**Locums usage assumptions.**")
        for item in editable_values:
            if item['cell'] in ['D17', 'B4']:
                label = f"{item['description']} ({item['cell']})"
                input_values[item['cell']] = st.slider(
                    label,
                    min_value=0,
                    max_value=int(item['value'] * 4),
                    value=int(item['value'])
                )
        locum_days_per_provider = st.slider(
        
            "Average Days per Locum per Month (Manual Entry)",
            min_value=0,
            max_value=240,
            value=20,
        key='Locum_Days_Slider'
    )
        input_values["Average Days per Locum per Month"] = locum_days_per_provider

    months = list(range(1, 25))

    permanent_onboard_rate = input_values.get('B21', 0)
    permanent_days_per_provider = input_values.get('B22', 0)
    permanent_open_days = input_values.get('C17', 0)
    float_pool_onboard_rate = input_values.get('B26', 0)
    float_pool_open_days = input_values.get('C17', 0)
    float_pool_days_per_provider = input_values.get('B27', 0)
    locum_open_days = input_values.get('D17', 0)
    hospitalist_rate = input_values.get('B4', 0)
    locum_days_per_provider = input_values.get('Average Days per Locum per Month', 0)

    permanent_shifts = []
    max_monthly_shifts = st.sidebar.slider(
    "Maximum Monthly Shifts (System Cap)",
    min_value=500,
    max_value=5000,
    value=1960,
    step=100,
    key='Max_Shift_Cap_Slider'
)
    total_permanent = 0
    for month in months:
        if month >= 4:
            total_permanent += permanent_onboard_rate
        permanent_shifts.append(min(total_permanent * permanent_days_per_provider, max_monthly_shifts))

    float_pool_shifts = []
    total_float_pool = 0
    for month in months:
        if month >= 12:
            total_float_pool += float_pool_onboard_rate
            float_pool_shifts.append(min(total_float_pool * float_pool_days_per_provider, max_monthly_shifts))
        else:
            float_pool_shifts.append(0)

    locum_shifts = []
    for idx, month in enumerate(months):
        if month >= 4:
            projected_coverage = permanent_shifts[idx] + float_pool_shifts[idx]
            locum_demand = max(max_monthly_shifts - projected_coverage, 0)
            decay_factor = max(locum_demand / max_monthly_shifts, 0)
            locum_shifts.append(int(locum_open_days * locum_days_per_provider * decay_factor))
        else:
            locum_shifts.append(0)

    shifts_data = {
        'Permanent': permanent_shifts,
        'Float Pool': float_pool_shifts,
        'VISTA Locums': locum_shifts
    }

    cost_data = {
        'Permanent': [shifts * 100 for shifts in permanent_shifts],
        'Float Pool': [shifts * 80 for shifts in float_pool_shifts],
        'VISTA Locums': [shifts * hospitalist_rate for shifts in locum_shifts]
    }

    total_cost_per_month = [
        cost_data['Permanent'][i] + cost_data['Float Pool'][i] + cost_data['VISTA Locums'][i]
        for i in range(len(months))
    ]

    baseline_monthly_cost = st.sidebar.slider(
    "Baseline Monthly Cost", min_value=100000, max_value=2000000, value=500000, step=50000,
    key='Baseline_Cost_Slider'
)
    total_baseline_cost = baseline_monthly_cost * len(months)
    total_actual_cost = sum(total_cost_per_month)
    total_savings = total_baseline_cost - total_actual_cost

    locum_providers = [int(shifts / locum_days_per_provider) if locum_days_per_provider > 0 else 0 for shifts in locum_shifts]
    total_locum_providers = sum(locum_providers)

    st.subheader("📈 Shift Coverage Over 24 Months")
    fig1 = go.Figure()
    for cat in ['Permanent', 'Float Pool', 'VISTA Locums']:
        fig1.add_trace(go.Bar(x=months, y=shifts_data[cat], name=cat, hovertemplate=f"%{{x}} Month<br>%{{y}} Shifts"))
    fig1.update_layout(barmode='stack', xaxis_title="Month", yaxis_title="Shifts")
    st.plotly_chart(fig1, use_container_width=True)

    st.subheader("💰 Staffing Costs & Savings Over 24 Months")
    fig2 = go.Figure()
    for cat in ['Permanent', 'Float Pool', 'VISTA Locums']:
        fig2.add_trace(go.Bar(x=months, y=cost_data[cat], name=cat, hovertemplate=f"%{{x}} Month<br>$%{{y:,.0f}} Cost"))
    fig2.add_trace(go.Scatter(x=months, y=total_cost_per_month, mode='lines+markers', name='Total Monthly Cost', line=dict(color='black', dash='dot')))
    fig2.update_layout(barmode='stack', xaxis_title="Month", yaxis_title="Cost ($)")
    st.plotly_chart(fig2, use_container_width=True)

    st.subheader("👨‍⚕️ Locum Providers Per Month")
    fig3 = go.Figure()
    fig3.add_trace(go.Scatter(x=months, y=locum_providers, mode='lines+markers', name='Locum Providers'))
    fig3.update_layout(xaxis_title="Month", yaxis_title="Locum Providers")
    st.plotly_chart(fig3, use_container_width=True)

    st.info(f"💰 **Total Cost Over 24 Months:** ${total_actual_cost:,.0f}")
    st.info(f"🏥 **Permanent Cost:** ${sum(cost_data['Permanent']):,.0f}")
    st.info(f"🧑‍⚕️ **Float Pool Cost:** ${sum(cost_data['Float Pool']):,.0f}")
    st.info(f"🩺 **VISTA Locums Cost:** ${sum(cost_data['VISTA Locums']):,.0f}")
    if total_savings >= 0:
        st.success(f"🎯 **Total Savings vs Baseline:** ${total_savings:,.0f}")
    else:
        st.error(f"⚠️ **Over Baseline by:** ${abs(total_savings):,.0f}")

    ".replace("
with tab2:
        st.subheader("🧮 Shift-Based ROI Calculator")
        st.markdown("Upload your Excel file with ICU Bed Loss data to calculate shift-based ROI.")

        uploaded_file_icu = st.file_uploader("Upload ICU Bed Loss Excel File", type=["xlsx"], key='ICU_File')

        if uploaded_file_icu is not None:
            df = pd.read_excel(uploaded_file_icu)

            st.sidebar.subheader("Adjust ICU Bed Loss Model Inputs")
            st.sidebar.markdown("---")
            locum_shift_cost = st.sidebar.number_input("Locum Shift Cost ($)", min_value=0, value=2000, step=100)
            icu_day_loss = st.sidebar.number_input("ICU Revenue Loss per Day ($)", min_value=0, value=10000, step=500)

            df['Total_Locum_Cost'] = df['Locum Shifts'] * locum_shift_cost
            df['Total_ICU_Loss'] = df['ICU Bed Loss Days'] * icu_day_loss
            df['Total_Impact'] = df['Total_Locum_Cost'] + df['Total_ICU_Loss']

            st.subheader("📊 ICU Bed Loss Impact Analysis")

            fig = go.Figure()
            fig.add_trace(go.Bar(x=df['Month'], y=df['Total_Locum_Cost'], name='Total Locum Cost'))
            fig.add_trace(go.Bar(x=df['Month'], y=df['Total_ICU_Loss'], name='Total ICU Revenue Loss'))
            fig.update_layout(barmode='stack', xaxis_title='Month', yaxis_title='Total Impact ($)')
            st.plotly_chart(fig, use_container_width=True)

            st.subheader("💰 Total Financial Impact")
            st.metric("Total Locum Cost", f"${df['Total_Locum_Cost'].sum():,.0f}")
            st.metric("Total ICU Revenue Loss", f"${df['Total_ICU_Loss'].sum():,.0f}")
            st.metric("Total Combined Impact", f"${df['Total_Impact'].sum():,.0f}")

        else:
            st.info("Please upload an ICU Bed Loss Excel file to use this calculator.")

else:
    st.info("Please upload an Excel file to get started.")
